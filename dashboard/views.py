from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count, F, Case, When, DecimalField
from django.utils import timezone
from datetime import datetime, timedelta, date
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Siparis, UserProfile, Notification, Transaction, TransactionCategory, Event, MalzemeHareketi, MalzemeDosya
from .forms import SiparisForm, TransactionForm, MalzemeExcelUploadForm
# pandas removed - using openpyxl instead
from collections import defaultdict
from django.views.decorators.csrf import csrf_exempt
from django.utils.dateparse import parse_date

@login_required
def index(request):
    """Dashboard ana sayfası"""
    # Sadece kontrol edilen siparişlerden lastik satış analizi verileri (sadece kullanıcının siparişleri)
    kontrol_siparisler = Siparis.objects.filter(durum='kontrol', user=request.user)
    
    # Lastik satış analizi verileri - mevsim ve araç tipi bazında
    tire_sales_data = {
        'yaz': {
            'binek': kontrol_siparisler.filter(mevsim='yaz', grup='binek').aggregate(total=Sum('adet'))['total'] or 0,
            'ticari': kontrol_siparisler.filter(mevsim='yaz', grup='ticari').aggregate(total=Sum('adet'))['total'] or 0
        },
        'kis': {
            'binek': kontrol_siparisler.filter(mevsim='kis', grup='binek').aggregate(total=Sum('adet'))['total'] or 0,
            'ticari': kontrol_siparisler.filter(mevsim='kis', grup='ticari').aggregate(total=Sum('adet'))['total'] or 0
        },
        'dort_mevsim': {
            'binek': kontrol_siparisler.filter(mevsim='dort-mevsim', grup='binek').aggregate(total=Sum('adet'))['total'] or 0,
            'ticari': kontrol_siparisler.filter(mevsim='dort-mevsim', grup='ticari').aggregate(total=Sum('adet'))['total'] or 0
        }
    }
    
    # Lastik marka dağılımı verileri - sadece kontrol edilen siparişler
    brand_distribution = (
        kontrol_siparisler
        .values('marka')
        .annotate(total_adet=Sum('adet'))
        .order_by('-total_adet')
    )
    
    # Chart için veri hazırlama
    brand_labels = []
    brand_data = []
    brand_colors = [
        '#3b82f6', '#ef4444', '#10b981', '#f59e0b', 
        '#8b5cf6', '#ec4899', '#06b6d4', '#84cc16'
    ]
    
    for i, brand in enumerate(brand_distribution):
        brand_labels.append(brand['marka'])
        brand_data.append(brand['total_adet'])
    
    # Real brand chart data from user's orders
    brand_chart_data = {
        'labels': brand_labels,
        'data': brand_data,
        'colors': brand_colors[:len(brand_labels)]
    }
    
    # Son eklenen işlemler (kullanıcının siparişlerinden son 5 kayıt)
    son_islemler = Siparis.objects.filter(user=request.user).order_by('-olusturma_tarihi')[:5]
    
    # Son finance işlemleri (kullanıcının işlemlerinden son 5 kayıt)
    son_finance_islemleri = Transaction.objects.filter(created_by=request.user).order_by('-created_at')[:5]
    
    # Gerçek istatistikler - sadece kullanıcının siparişleri
    toplam_siparis = Siparis.objects.filter(user=request.user).count()
    toplam_ciro = Siparis.objects.filter(user=request.user).aggregate(total=Sum('toplam_fiyat'))['total'] or 0
    kontrol_edilen_siparis = kontrol_siparisler.count()
    toplam_adet = Siparis.objects.filter(user=request.user).aggregate(total=Sum('adet'))['total'] or 0
    
    # Aylık gelir/gider verileri (son 12 ay) - Transaction üzerinden (kullanıcıya göre)
    from datetime import datetime, timedelta
    monthly_income = []
    monthly_expense = []
    monthly_labels = []
    toplam_ifade = (F('nakit') + F('kredi_karti') + F('cari') + F('mehmet_havale'))
    
    for i in range(11, -1, -1):
        start_date = timezone.now() - timedelta(days=30*i)
        end_date = start_date + timedelta(days=30)
        
        gelir_toplam = Transaction.objects.filter(
            created_by=request.user,
            tarih__gte=start_date.date(),
            tarih__lt=end_date.date(),
            hareket_tipi='gelir'
        ).aggregate(total=Sum(toplam_ifade))['total'] or 0
        
        gider_toplam = Transaction.objects.filter(
            created_by=request.user,
            tarih__gte=start_date.date(),
            tarih__lt=end_date.date(),
            hareket_tipi='gider'
        ).aggregate(total=Sum(toplam_ifade))['total'] or 0
        
        monthly_income.append(float(gelir_toplam))
        monthly_expense.append(float(gider_toplam))
        monthly_labels.append(start_date.strftime('%b'))
    
    # En çok alım yaptığımız cariler (fiyat bazında) - Yeni Sipariş Ekle'den oluşturulan veriler
    # Sadece kontrol edilmiş siparişlerden en çok alım yapan cariler (sadece kullanıcının siparişleri)
    top_customers = (
        Siparis.objects
        .filter(durum='kontrol', user=request.user)  # Sadece kontrol edilmiş siparişler ve kullanıcının siparişleri
        .values('cari_firma')
        .annotate(
            total_purchase=Sum('toplam_fiyat'),
            siparis_sayisi=Count('id')
        )
        .order_by('-total_purchase')[:8]
    )
    
    customer_labels = []
    customer_data = []
    customer_details = []
    
    for customer in top_customers:
        # Firma adını kısalt (çok uzunsa)
        firma_name = customer['cari_firma']
        original_name = firma_name
        if len(firma_name) > 20:
            firma_name = firma_name[:17] + '...'
        
        customer_labels.append(firma_name)
        customer_data.append(float(customer['total_purchase']))
        customer_details.append({
            'name': original_name,
            'total': float(customer['total_purchase']),
            'count': customer['siparis_sayisi']
        })
    
    context = {
        'page_title': 'Dashboard',
        'stats': {
            'total_users': toplam_siparis,
            'revenue': toplam_ciro,
            'orders': kontrol_edilen_siparis,
            'avg_response': toplam_adet
        },
        'tire_sales_data': json.dumps(tire_sales_data),
        'brand_chart_data': json.dumps(brand_chart_data),
        'son_islemler': son_islemler,
        'son_finance_islemleri': son_finance_islemleri,
        'monthly_income': json.dumps(monthly_income),
        'monthly_expense': json.dumps(monthly_expense),
        'monthly_labels': json.dumps(monthly_labels),
        'top_customers_data': json.dumps({
            'labels': customer_labels,
            'data': customer_data,
            'details': customer_details
        })
    }
    return render(request, 'dashboard/index.html', context)

def analytics(request):
    """Analytics sayfası"""
    context = {
        'page_title': 'Analytics',
    }
    return render(request, 'dashboard/analytics.html', context)

@login_required
def users(request):
    """Users sayfası - Rol bazlı yetkilendirme"""
    # Kullanıcının profilini al
    try:
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        # Eğer profil yoksa varsayılan olarak yönetici rolü ver
        user_profile = UserProfile.objects.create(user=request.user, role='yonetici')
    
    # Admin kontrolü
    is_admin = user_profile.is_admin()
    
    # Kullanıcı listesi - sadece admin görebilir
    users_list = []
    if is_admin:
        users_list = User.objects.all().select_related('userprofile')
    
    # Kullanıcı oluşturma
    if request.method == 'POST' and is_admin:
        username = request.POST.get('username')
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        password = request.POST.get('password')
        role = request.POST.get('role', 'yonetici')
        
        if username and password:
            try:
                # Kullanıcı oluştur
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name
                )
                
                # Profil oluştur
                UserProfile.objects.create(user=user, role=role)
                
                messages.success(request, f'Kullanıcı {username} başarıyla oluşturuldu!')
                return redirect('dashboard:users')
            except Exception as e:
                messages.error(request, f'Kullanıcı oluşturulurken hata: {str(e)}')
        else:
            messages.error(request, 'Kullanıcı adı ve şifre gereklidir!')
    
    context = {
        'page_title': 'Users',
        'is_admin': is_admin,
        'users_list': users_list,
        'user_profile': user_profile,
    }
    return render(request, 'dashboard/users.html', context)

@login_required
def products(request):
    """Products sayfası: tarih aralığı filtresi ile işlemleri göster"""
    from datetime import date
    
    # Tarih aralığı parametrelerini al
    baslangic_tarih = request.GET.get('baslangic_tarih')
    bitis_tarih = request.GET.get('bitis_tarih')
    
    # Eski tek tarih parametresi (geriye uyumluluk için)
    secilen_tarih = request.GET.get('tarih')
    if not secilen_tarih and not baslangic_tarih and not bitis_tarih:
        secilen_tarih = date.today().strftime('%Y-%m-%d')
    
    # İşlemleri filtrele
    qs = Transaction.objects.filter(created_by=request.user)
    
    if baslangic_tarih:
        qs = qs.filter(tarih__gte=baslangic_tarih)
    if bitis_tarih:
        qs = qs.filter(tarih__lte=bitis_tarih)
    if secilen_tarih and not baslangic_tarih and not bitis_tarih:
        qs = qs.filter(tarih=secilen_tarih)
    
    qs = qs.order_by('-created_at')
    
    # Özet bilgileri
    toplam_ifade = (F('nakit') + F('kredi_karti') + F('cari') + F('mehmet_havale'))
    
    gun_ozeti = qs.aggregate(
        gelir=Sum(Case(When(hareket_tipi='gelir', then=toplam_ifade), default=0, output_field=DecimalField(max_digits=12, decimal_places=2))),
        gider=Sum(Case(When(hareket_tipi='gider', then=toplam_ifade), default=0, output_field=DecimalField(max_digits=12, decimal_places=2)))
    )
    
    gun_ozeti['gelir'] = gun_ozeti['gelir'] or 0
    gun_ozeti['gider'] = gun_ozeti['gider'] or 0
    gun_ozeti['net'] = gun_ozeti['gelir'] - gun_ozeti['gider']
    gun_ozeti['islem_sayisi'] = qs.count()
    
    # Ödeme yöntemlerine göre toplamlar (gelir - gider)
    # Gelir toplamları
    gelir_nakit = qs.filter(hareket_tipi='gelir').aggregate(total=Sum('nakit', default=0))['total'] or 0
    gelir_kredi_karti = qs.filter(hareket_tipi='gelir').aggregate(total=Sum('kredi_karti', default=0))['total'] or 0
    gelir_cari = qs.filter(hareket_tipi='gelir').aggregate(total=Sum('cari', default=0))['total'] or 0
    gelir_mehmet_havale = qs.filter(hareket_tipi='gelir').aggregate(total=Sum('mehmet_havale', default=0))['total'] or 0
    
    # Gider toplamları
    gider_nakit = qs.filter(hareket_tipi='gider').aggregate(total=Sum('nakit', default=0))['total'] or 0
    gider_kredi_karti = qs.filter(hareket_tipi='gider').aggregate(total=Sum('kredi_karti', default=0))['total'] or 0
    gider_cari = qs.filter(hareket_tipi='gider').aggregate(total=Sum('cari', default=0))['total'] or 0
    gider_mehmet_havale = qs.filter(hareket_tipi='gider').aggregate(total=Sum('mehmet_havale', default=0))['total'] or 0
    
    # Net toplamlar (gelir - gider)
    gun_ozeti['nakit_toplam'] = gelir_nakit - gider_nakit
    gun_ozeti['kredi_karti_toplam'] = gelir_kredi_karti - gider_kredi_karti
    gun_ozeti['cari_toplam'] = gelir_cari - gider_cari
    gun_ozeti['mehmet_havale_toplam'] = gelir_mehmet_havale - gider_mehmet_havale

    # Excel verileri için tarih filtrelemesi
    hareketler = MalzemeHareketi.objects.filter(kullanici=request.user).order_by('-tarih')
    dosyalar = MalzemeDosya.objects.filter(kullanici=request.user).prefetch_related('satirlar').order_by('-yukleme_tarihi')
    
    # Tarih aralığı filtresi uygula
    if baslangic_tarih:
        hareketler = hareketler.filter(tarih__gte=baslangic_tarih)
        dosyalar = dosyalar.filter(yukleme_tarihi__date__gte=baslangic_tarih)
    if bitis_tarih:
        hareketler = hareketler.filter(tarih__lte=bitis_tarih)
        dosyalar = dosyalar.filter(yukleme_tarihi__date__lte=bitis_tarih)
    
    # Eski parametreler (geriye uyumluluk)
    start_s = request.GET.get('start-date')
    end_s = request.GET.get('end-date')
    if start_s and not baslangic_tarih:
        hareketler = hareketler.filter(tarih__gte=start_s)
        dosyalar = dosyalar.filter(yukleme_tarihi__date__gte=start_s)
    if end_s and not bitis_tarih:
        hareketler = hareketler.filter(tarih__lte=end_s)
        dosyalar = dosyalar.filter(yukleme_tarihi__date__lte=end_s)
    from collections import defaultdict
    import datetime
    gunluk = defaultdict(list)
    for h in hareketler:
        tarih_str = h.tarih.strftime('%d.%m.%Y')
        gunluk[tarih_str].append(h)
    gunluk_excel = defaultdict(list)
    for d in dosyalar:
        # Dosyayı hem yükleme tarihine hem de içindeki kayıtların tarihlerine göre gruplandır
        yukleme_tarihi_str = d.yukleme_tarihi.strftime('%d.%m.%Y')
        gunluk_excel[yukleme_tarihi_str].append(d)
        
        # Ayrıca dosya içindeki kayıtların tarihlerine göre de gruplandır
        for satir in d.satirlar.all():
            satir_tarih_str = satir.tarih.strftime('%d.%m.%Y')
            if satir_tarih_str != yukleme_tarihi_str:
                # Eğer kayıt tarihi farklıysa, o tarihe de ekle
                if d not in gunluk_excel[satir_tarih_str]:
                    gunluk_excel[satir_tarih_str].append(d)
    gunluk_sorted = dict(sorted(gunluk.items(), key=lambda x: datetime.datetime.strptime(x[0], '%d.%m.%Y'), reverse=True))
    gunluk_excel_sorted = dict(sorted(gunluk_excel.items(), key=lambda x: datetime.datetime.strptime(x[0], '%d.%m.%Y'), reverse=True))
    
    # Excel verilerini belirle
    secilen_gun_excel = []
    secilen_tarih_str = ''
    
    if secilen_tarih and not baslangic_tarih and not bitis_tarih:
        # Tek tarih seçilmişse
        from datetime import datetime as dt
        secilen_tarih_obj = dt.strptime(secilen_tarih, '%Y-%m-%d')
        secilen_tarih_str = secilen_tarih_obj.strftime('%d.%m.%Y')
        secilen_gun_excel = gunluk_excel_sorted.get(secilen_tarih_str, [])
    else:
        # Tarih aralığı seçilmişse, tüm filtrelenmiş dosyaları göster
        secilen_gun_excel = list(dosyalar)
    
    # Excel verilerine göre ödeme şekillerine göre toplamlar
    excel_odeme_toplamlari = {}
    
    # Tüm filtrelenmiş Excel satırlarını al
    excel_satirlar = MalzemeHareketi.objects.filter(kullanici=request.user)
    if baslangic_tarih:
        excel_satirlar = excel_satirlar.filter(tarih__gte=baslangic_tarih)
    if bitis_tarih:
        excel_satirlar = excel_satirlar.filter(tarih__lte=bitis_tarih)
    if secilen_tarih and not baslangic_tarih and not bitis_tarih:
        excel_satirlar = excel_satirlar.filter(tarih=secilen_tarih)
    
    # Ödeme şekillerine göre grupla ve topla
    excel_odeme_dict = {
        'Nakit': 0,
        'Kredi_Karti': 0,
        'Cari': 0,
        'Sanal_Pos': 0
    }
    
    for satir in excel_satirlar:
        odeme_sekli = satir.odeme_sekli or 'Belirtilmemiş'
        # Ödeme şeklini normalize et (küçük harfe çevir, boşlukları temizle)
        odeme_sekli_normalized = odeme_sekli.lower().strip().replace(' ', '').replace('.', '').replace('-', '')
        odeme_sekli_original = odeme_sekli.strip()
        
        # Ana ödeme şekillerini belirle (tam eşleşme ve içerme kontrolü)
        # Nakit kontrolü
        if (odeme_sekli_original.lower() == 'nakit' or 
            'nakit' in odeme_sekli_normalized or 
            odeme_sekli_normalized == 'nakit'):
            excel_odeme_dict['Nakit'] += float(satir.tutar)
        # Kredi Kartı kontrolü
        elif ('kart' in odeme_sekli_normalized or 
              'kredi' in odeme_sekli_normalized or 
              'kredit' in odeme_sekli_normalized or
              odeme_sekli_original.lower() in ['kredi kartı', 'kredi karti', 'kart', 'credit card']):
            excel_odeme_dict['Kredi_Karti'] += float(satir.tutar)
        # Sanal Pos kontrolü
        elif ('sanal' in odeme_sekli_normalized or 
              'pos' in odeme_sekli_normalized or
              'sanalpos' in odeme_sekli_normalized or
              'sanal pos' in odeme_sekli_original.lower() or
              'havale' in odeme_sekli_normalized or 
              'mehmet' in odeme_sekli_normalized or
              'mhavale' in odeme_sekli_normalized or
              odeme_sekli_original.lower() in ['sanal pos', 'sanalpos', 'pos', 'm.havale', 'm havale', 'mhavale', 'mehmet havale']):
            excel_odeme_dict['Sanal_Pos'] += float(satir.tutar)
        # Cari kontrolü
        elif ('cari' in odeme_sekli_normalized or 
              odeme_sekli_original.lower() == 'cari' or
              odeme_sekli_normalized == 'cari'):
            excel_odeme_dict['Cari'] += float(satir.tutar)
        else:
            # Eğer eşleşme yoksa, varsayılan olarak hiçbir şeye eklenmez
            pass
    
    # Dictionary'yi context'e gönder
    excel_odeme_toplamlari = excel_odeme_dict
    
    context = {
        'page_title': 'Products',
        'secilen_tarih': secilen_tarih or date.today().strftime('%Y-%m-%d'),
        'baslangic_tarih': baslangic_tarih,
        'bitis_tarih': bitis_tarih,
        'gun_ozeti': gun_ozeti,
        'gunun_islemleri': qs,
        'malzeme_gunluk': gunluk_sorted,
        'gunluk_excel': gunluk_excel_sorted,
        'secilen_gun_excel': secilen_gun_excel,
        'start_selected': start_s or '',
        'end_selected': end_s or '',
        # Debug bilgileri
        'debug_secilen_tarih_str': secilen_tarih_str,
        'debug_gunluk_excel_keys': list(gunluk_excel_sorted.keys()),
        'debug_toplam_dosya': len(dosyalar),
        'excel_odeme_toplamlari': excel_odeme_toplamlari,
    }
    return render(request, 'dashboard/products.html', context)

@login_required
def finance(request):
    """Gelir/Gider İşlemleri form sayfası"""
    form = TransactionForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            tx = form.save(commit=False)
            tx.created_by = request.user
            tx.save()
            messages.success(request, 'İşlem kaydedildi.')
            next_url = request.POST.get('next') or request.GET.get('next')
            if next_url and next_url.startswith('/'):
                return redirect(next_url)
            return redirect('dashboard:products')
        else:
            # Form geçersizse hataları göster
            messages.error(request, 'Form hatası! Lütfen aşağıdaki hataları kontrol edin.')
            print("Form Errors:", form.errors)  # Debug için

    ana_kategoriler = TransactionCategory.objects.filter(parent=None).order_by('name')
    last_transactions = Transaction.objects.filter(created_by=request.user).order_by('-tarih', '-id')[:10]
    
    # Debug için
    print(f"Ana Kategoriler Sayısı: {ana_kategoriler.count()}")
    for kat in ana_kategoriler:
        print(f"  - {kat.name} (ID: {kat.id})")

    context = {
        'page_title': 'Gelir/Gider İşlemleri',
        'form': form,
        'ana_kategoriler': ana_kategoriler,
        'transactions': last_transactions,
    }
    return render(request, 'dashboard/finance.html', context)

@login_required
def orders(request):
    """Sipariş Envanteri Dashboard (Orders)"""
    # Hızlı arama (cari ile arama)
    query = request.GET.get('q', '').strip()

    siparisler = Siparis.objects.filter(user=request.user)
    if query:
        siparisler = siparisler.filter(cari_firma__icontains=query)

    # Metrikler
    toplam_adet = siparisler.aggregate(total=Sum('adet'))['total'] or 0
    stoktaki_adet = siparisler.filter(ambar='stok').aggregate(total=Sum('adet'))['total'] or 0
    satistaki_adet = siparisler.filter(ambar='satis').aggregate(total=Sum('adet'))['total'] or 0
    toplam_ciro = siparisler.aggregate(total=Sum('toplam_fiyat'))['total'] or 0
    yolda_siparisler = siparisler.filter(durum='yolda').count()
    islemde_siparisler = siparisler.filter(durum='islemde').count()

    # Son eklenen işlemler (son 10 kayıt)
    son_islemler = siparisler.order_by('-olusturma_tarihi')[:10]

    # Marka dağılımı (marka bazında adet toplamı ve kayıt sayısı)
    marka_dagilimi = (
        siparisler.values('marka')
        .annotate(
            kayit_sayisi=Count('id'),
            adet_toplam=Sum('adet'),
            tutar_toplam=Sum('toplam_fiyat'),
        )
        .order_by('-adet_toplam')
    )

    context = {
        'page_title': 'Sipariş Envanteri Dashboard',
        'q': query,
        'stats': {
            'toplam_adet': toplam_adet,
            'stoktaki_adet': stoktaki_adet,
            'satistaki_adet': satistaki_adet,
            'toplam_ciro': toplam_ciro,
            'yolda': yolda_siparisler,
            'islemde': islemde_siparisler,
        },
        'son_islemler': son_islemler,
        'marka_dagilimi': marka_dagilimi,
    }
    return render(request, 'dashboard/orders.html', context)

@login_required
def forms(request):
    """Kontrol Edilen Siparişler Sayfası"""
    # Filtreleme parametreleri
    firma = request.GET.get('firma', '')
    marka = request.GET.get('marka', '')
    grup = request.GET.get('grup', '')
    mevsim = request.GET.get('mevsim', '')
    ambar = request.GET.get('ambar', '')
    tarih_filtre = request.GET.get('tarih', '')
    baslangic_tarihi = request.GET.get('baslangic_tarihi', '')
    bitis_tarihi = request.GET.get('bitis_tarihi', '')
    
    # Sadece kontrol edilen siparişleri getir (sadece kullanıcının siparişleri)
    siparisler = Siparis.objects.filter(durum='kontrol', user=request.user)
    
    # Filtreleme uygula
    if firma:
        siparisler = siparisler.filter(cari_firma__icontains=firma)
    if marka:
        siparisler = siparisler.filter(marka__icontains=marka)
    if grup:
        siparisler = siparisler.filter(grup=grup)
    if mevsim:
        siparisler = siparisler.filter(mevsim=mevsim)
    if ambar:
        siparisler = siparisler.filter(ambar=ambar)
    
    # Tarih filtreleme uygula
    now = timezone.now()
    if tarih_filtre:
        if tarih_filtre == 'son-1-ay':
            start_date = now - timedelta(days=30)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'son-3-ay':
            start_date = now - timedelta(days=90)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'son-6-ay':
            start_date = now - timedelta(days=180)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'bugun':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = now.replace(hour=23, minute=59, second=59, microsecond=999999)
            siparisler = siparisler.filter(olusturma_tarihi__range=[start_date, end_date])
        elif tarih_filtre == 'bu-hafta':
            start_date = now - timedelta(days=now.weekday())
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'bu-ay':
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
    
    # Özel tarih aralığı filtreleme
    if baslangic_tarihi and bitis_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0))
            end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
            siparisler = siparisler.filter(olusturma_tarihi__range=[start_date, end_date])
        except ValueError:
            pass
    elif baslangic_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0))
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        except ValueError:
            pass
    elif bitis_tarihi:
        try:
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d')
            end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
            siparisler = siparisler.filter(olusturma_tarihi__lte=end_date)
        except ValueError:
            pass
    
    # İstatistikler hesapla
    toplam_kontrol = siparisler.count()
    toplam_tutar = siparisler.aggregate(total=Sum('toplam_fiyat'))['total'] or 0
    toplam_adet = siparisler.aggregate(total=Sum('adet'))['total'] or 0
    
    # Grup bazında kontrol istatistikleri (lastik adet toplamı)
    grup_istatistikleri = siparisler.values('grup').annotate(
        total_adet=Sum('adet'),
        total_amount=Sum('toplam_fiyat')
    ).order_by('-total_adet')
    
    
    # Sayfalama
    paginator = Paginator(siparisler, 15)  # Sayfa başına 15 kayıt
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_title': 'Kontrol Edilen Siparişler',
        'siparisler': page_obj,
        'filters': {
            'firma': firma,
            'marka': marka,
            'grup': grup,
            'mevsim': mevsim,
            'ambar': ambar,
            'tarih': tarih_filtre,
            'baslangic_tarihi': baslangic_tarihi,
            'bitis_tarihi': bitis_tarihi,
        },
        'stats': {
            'toplam_kontrol': toplam_kontrol,
            'toplam_tutar': toplam_tutar,
            'toplam_adet': toplam_adet,
        },
        'grup_istatistikleri': grup_istatistikleri,
    }
    return render(request, 'dashboard/forms.html', context)

@login_required
def elements(request):
    """Sipariş Envanteri Listesi sayfası"""
    # Filtreleme parametreleri
    firma = request.GET.get('firma', '')
    marka = request.GET.get('marka', '')
    grup = request.GET.get('grup', '')
    durum = request.GET.get('durum', '')
    mevsim = request.GET.get('mevsim', '')
    ambar = request.GET.get('ambar', '')
    tarih_filtre = request.GET.get('tarih', '')
    baslangic_tarihi = request.GET.get('baslangic_tarihi', '')
    bitis_tarihi = request.GET.get('bitis_tarihi', '')
    
    # Siparişleri getir (iptal edilenleri ve kontrol edilenleri hariç tut, sadece kullanıcının siparişleri)
    siparisler = Siparis.objects.filter(user=request.user).exclude(durum__in=['iptal', 'kontrol'])
    
    # Filtreleme uygula
    if firma:
        siparisler = siparisler.filter(cari_firma__icontains=firma)
    if marka:
        siparisler = siparisler.filter(marka__icontains=marka)
    if grup:
        siparisler = siparisler.filter(grup=grup)
    if durum:
        siparisler = siparisler.filter(durum=durum)
    if mevsim:
        siparisler = siparisler.filter(mevsim=mevsim)
    if ambar:
        siparisler = siparisler.filter(ambar=ambar)
    
    # Tarih filtreleme uygula
    now = timezone.now()
    if tarih_filtre:
        if tarih_filtre == 'son-1-ay':
            start_date = now - timedelta(days=30)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'son-3-ay':
            start_date = now - timedelta(days=90)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'son-6-ay':
            start_date = now - timedelta(days=180)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'bugun':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = now.replace(hour=23, minute=59, second=59, microsecond=999999)
            siparisler = siparisler.filter(olusturma_tarihi__range=[start_date, end_date])
        elif tarih_filtre == 'bu-hafta':
            start_date = now - timedelta(days=now.weekday())
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'bu-ay':
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
    
    # Özel tarih aralığı filtreleme
    if baslangic_tarihi and bitis_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d')
            # Tarih aralığını günün başı ve sonu olarak ayarla
            start_date = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0))
            end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
            siparisler = siparisler.filter(olusturma_tarihi__range=[start_date, end_date])
        except ValueError:
            # Geçersiz tarih formatı
            pass
    elif baslangic_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0))
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        except ValueError:
            pass
    elif bitis_tarihi:
        try:
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d')
            end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
            siparisler = siparisler.filter(olusturma_tarihi__lte=end_date)
        except ValueError:
            pass
    
    # Sayfalama
    paginator = Paginator(siparisler, 10)  # Sayfa başına 10 kayıt
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_title': 'Sipariş Envanteri',
        'siparisler': page_obj,
        'filters': {
            'firma': firma,
            'marka': marka,
            'grup': grup,
            'durum': durum,
            'mevsim': mevsim,
            'ambar': ambar,
            'tarih': tarih_filtre,
            'baslangic_tarihi': baslangic_tarihi,
            'bitis_tarihi': bitis_tarihi,
        }
    }
    return render(request, 'dashboard/elements.html', context)

def elements_buttons(request):
    """Elements buttons sayfası"""
    context = {
        'page_title': 'Elements - Buttons',
    }
    return render(request, 'dashboard/elements-buttons.html', context)

def elements_alerts(request):
    """Elements alerts sayfası"""
    context = {
        'page_title': 'Elements - Alerts',
    }
    return render(request, 'dashboard/elements-alerts.html', context)

def elements_badges(request):
    """Elements badges sayfası"""
    context = {
        'page_title': 'Elements - Badges',
    }
    return render(request, 'dashboard/elements-badges.html', context)

def elements_cards(request):
    """Elements cards sayfası"""
    context = {
        'page_title': 'Elements - Cards',
    }
    return render(request, 'dashboard/elements-cards.html', context)

def elements_modals(request):
    """Elements modals sayfası"""
    context = {
        'page_title': 'Elements - Modals',
    }
    return render(request, 'dashboard/elements-modals.html', context)

def elements_forms(request):
    """Elements forms sayfası"""
    context = {
        'page_title': 'Elements - Forms',
    }
    return render(request, 'dashboard/elements-forms.html', context)

def elements_tables(request):
    """Elements tables sayfası"""
    context = {
        'page_title': 'Elements - Tables',
    }
    return render(request, 'dashboard/elements-tables.html', context)

@login_required
def yeni_lastik(request):
    """Yeni lastik ekleme sayfası"""
    if request.method == 'POST':
        form = SiparisForm(request.POST)
        if form.is_valid():
            siparis = form.save(commit=False)
            # Kullanıcıyı otomatik ata
            siparis.user = request.user
            siparis.save()
            messages.success(request, f'Sipariş başarıyla kaydedildi! ID: {siparis.id}')
            return redirect('dashboard:elements')
        else:
            # Form hatalarını detaylı göster
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f'{field}: {error}')
            messages.error(request, f'Form hataları: {", ".join(error_messages)}')
            print(f"Form errors: {form.errors}")  # Debug için
    else:
        form = SiparisForm()
    
    context = {
        'page_title': 'Yeni Lastik Ekle',
        'form': form,
    }
    return render(request, 'dashboard/yeni_lastik.html', context)

@login_required
def siparis_detay(request, siparis_id):
    """Sipariş detay sayfası"""
    siparis = get_object_or_404(Siparis, id=siparis_id, user=request.user)
    context = {
        'page_title': f'Sipariş Detayı - #{siparis.id}',
        'siparis': siparis,
    }
    return render(request, 'dashboard/siparis_detay.html', context)

@login_required
def siparis_duzenle(request, siparis_id):
    """Sipariş düzenleme sayfası"""
    siparis = get_object_or_404(Siparis, id=siparis_id, user=request.user)
    
    if request.method == 'POST':
        form = SiparisForm(request.POST, instance=siparis)
        if form.is_valid():
            form.save()
            messages.success(request, f'Sipariş #{siparis.id} başarıyla güncellendi!')
            return redirect('dashboard:elements')
        else:
            messages.error(request, 'Form hataları var. Lütfen kontrol edin.')
    else:
        form = SiparisForm(instance=siparis)
    
    context = {
        'page_title': f'Sipariş Düzenle - #{siparis.id}',
        'form': form,
        'siparis': siparis,
    }
    return render(request, 'dashboard/siparis_duzenle.html', context)

@login_required
def siparis_sil(request, siparis_id):
    """Sipariş silme"""
    siparis = get_object_or_404(Siparis, id=siparis_id, user=request.user)
    
    if request.method == 'POST':
        siparis_id = siparis.id
        siparis.delete()
        messages.success(request, f'Sipariş #{siparis_id} başarıyla silindi!')
        return redirect('dashboard:elements')
    
    context = {
        'page_title': f'Sipariş Sil - #{siparis.id}',
        'siparis': siparis,
    }
    return render(request, 'dashboard/siparis_sil.html', context)

@login_required
def siparis_whatsapp(request, siparis_id):
    """WhatsApp mesajı gönder"""
    siparis = get_object_or_404(Siparis, id=siparis_id, user=request.user)
    
    # WhatsApp mesajı oluştur
    mesaj = f"""*Lastik Envanteri - {siparis.cari_firma}*

*Ürün:* {siparis.urun}
*Marka:* {siparis.marka}
*Adet:* {siparis.adet}
*Durum:* {siparis.get_durum_display()}
*Güncellenen Son Tarih:* {timezone.localtime(siparis.guncelleme_tarihi).strftime('%d.%m.%Y %H:%M')}"""
    
    # WhatsApp URL'si oluştur (telefon numarası placeholder)
    encoded_mesaj = mesaj.replace(' ', '%20').replace('\n', '%0A')
    whatsapp_url = f"https://wa.me/?text={encoded_mesaj}"
    
    return redirect(whatsapp_url)

@login_required
def reports(request):
    """İptal Edilen Siparişler Raporu"""
    # Filtreleme parametreleri
    firma = request.GET.get('firma', '')
    marka = request.GET.get('marka', '')
    grup = request.GET.get('grup', '')
    mevsim = request.GET.get('mevsim', '')
    ambar = request.GET.get('ambar', '')
    tarih_filtre = request.GET.get('tarih', '')
    baslangic_tarihi = request.GET.get('baslangic_tarihi', '')
    bitis_tarihi = request.GET.get('bitis_tarihi', '')
    
    # Sadece iptal edilen siparişleri getir (sadece kullanıcının siparişleri)
    siparisler = Siparis.objects.filter(durum='iptal', user=request.user)
    
    # Filtreleme uygula
    if firma:
        siparisler = siparisler.filter(cari_firma__icontains=firma)
    if marka:
        siparisler = siparisler.filter(marka__icontains=marka)
    if grup:
        siparisler = siparisler.filter(grup=grup)
    if mevsim:
        siparisler = siparisler.filter(mevsim=mevsim)
    if ambar:
        siparisler = siparisler.filter(ambar=ambar)
    
    # Tarih filtreleme uygula
    now = timezone.now()
    if tarih_filtre:
        if tarih_filtre == 'son-1-ay':
            start_date = now - timedelta(days=30)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'son-3-ay':
            start_date = now - timedelta(days=90)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'son-6-ay':
            start_date = now - timedelta(days=180)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'bugun':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = now.replace(hour=23, minute=59, second=59, microsecond=999999)
            siparisler = siparisler.filter(olusturma_tarihi__range=[start_date, end_date])
        elif tarih_filtre == 'bu-hafta':
            start_date = now - timedelta(days=now.weekday())
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'bu-ay':
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
    
    # Özel tarih aralığı filtreleme
    if baslangic_tarihi and bitis_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0))
            end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
            siparisler = siparisler.filter(olusturma_tarihi__range=[start_date, end_date])
        except ValueError:
            pass
    elif baslangic_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0))
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        except ValueError:
            pass
    elif bitis_tarihi:
        try:
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d')
            end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
            siparisler = siparisler.filter(olusturma_tarihi__lte=end_date)
        except ValueError:
            pass
    
    # İstatistikler hesapla
    toplam_iptal = siparisler.count()
    toplam_tutar = siparisler.aggregate(total=Sum('toplam_fiyat'))['total'] or 0
    toplam_adet = siparisler.aggregate(total=Sum('adet'))['total'] or 0
    
    # Grup bazında iptal istatistikleri (lastik adet toplamı)
    grup_istatistikleri = siparisler.values('grup').annotate(
        total_adet=Sum('adet'),
        total_amount=Sum('toplam_fiyat')
    ).order_by('-total_adet')
    
    # Sayfalama
    paginator = Paginator(siparisler, 15)  # Sayfa başına 15 kayıt
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_title': 'İptal Edilen Siparişler Raporu',
        'siparisler': page_obj,
        'filters': {
            'firma': firma,
            'marka': marka,
            'grup': grup,
            'mevsim': mevsim,
            'ambar': ambar,
            'tarih': tarih_filtre,
            'baslangic_tarihi': baslangic_tarihi,
            'bitis_tarihi': bitis_tarihi,
        },
        'stats': {
            'toplam_iptal': toplam_iptal,
            'toplam_tutar': toplam_tutar,
            'toplam_adet': toplam_adet,
        },
        'grup_istatistikleri': grup_istatistikleri,
    }
    return render(request, 'dashboard/reports.html', context)

def messages_view(request):
    """Gelir/Gider Toplamları sayfası (kasa bazında ve genel toplam)."""
    qs = Transaction.objects.filter(created_by=request.user)

    toplam_ifade = (F('nakit') + F('kredi_karti') + F('cari') + F('mehmet_havale'))

    # Kasa bazında gelir, gider ve net (gelir - gider)
    kasa_ozet = (
        qs.values('kasa_adi')
        .annotate(
            gelir=Sum(Case(When(hareket_tipi='gelir', then=toplam_ifade), default=0, output_field=DecimalField(max_digits=14, decimal_places=2))),
            gider=Sum(Case(When(hareket_tipi='gider', then=toplam_ifade), default=0, output_field=DecimalField(max_digits=14, decimal_places=2)))
        )
        .annotate(net=F('gelir') - F('gider'))
        .order_by('kasa_adi')
    )

    # Kartlar için belirli kasalar
    servis_toplam = 0
    merkez_satis_toplam = 0
    canta_toplam = 0
    mehmet_havale_toplam = 0
    genel_toplam = 0
    rows = []
    
    for row in kasa_ozet:
        kasa = row['kasa_adi']
        net = float(row['net'] or 0)
        rows.append({'kasa_adi': kasa, 'bakiye': net})
        genel_toplam += net
        
        if kasa == 'servis':
            servis_toplam = net
        elif kasa == 'merkez-satis':
            merkez_satis_toplam = net
        elif kasa == 'canta':
            canta_toplam = net
        elif kasa == 'mehmet-havale':
            mehmet_havale_toplam = net

    # Ayrıca mehmet_havale field'ından da toplam hesapla
    mehmet_havale_field_toplam = qs.aggregate(
        mehmet_havale_gelir=Sum(Case(When(hareket_tipi='gelir', then='mehmet_havale'), default=0, output_field=DecimalField(max_digits=14, decimal_places=2))),
        mehmet_havale_gider=Sum(Case(When(hareket_tipi='gider', then='mehmet_havale'), default=0, output_field=DecimalField(max_digits=14, decimal_places=2)))
    )
    
    # Mehmet havale field toplamını da ekle
    mehmet_havale_net = float((mehmet_havale_field_toplam['mehmet_havale_gelir'] or 0) - (mehmet_havale_field_toplam['mehmet_havale_gider'] or 0))
    if mehmet_havale_toplam == 0:  # Eğer kasa olarak mehmet-havale yoksa field toplamını kullan
        mehmet_havale_toplam = mehmet_havale_net
    
    # Ödeme yöntemlerine göre toplamlar (Detaylı İşlemler'den)
    # Gelir toplamları
    gelir_nakit = qs.filter(hareket_tipi='gelir').aggregate(total=Sum('nakit', default=0))['total'] or 0
    gelir_kredi_karti = qs.filter(hareket_tipi='gelir').aggregate(total=Sum('kredi_karti', default=0))['total'] or 0
    gelir_cari = qs.filter(hareket_tipi='gelir').aggregate(total=Sum('cari', default=0))['total'] or 0
    gelir_mehmet_havale = qs.filter(hareket_tipi='gelir').aggregate(total=Sum('mehmet_havale', default=0))['total'] or 0
    
    # Gider toplamları
    gider_nakit = qs.filter(hareket_tipi='gider').aggregate(total=Sum('nakit', default=0))['total'] or 0
    gider_kredi_karti = qs.filter(hareket_tipi='gider').aggregate(total=Sum('kredi_karti', default=0))['total'] or 0
    gider_cari = qs.filter(hareket_tipi='gider').aggregate(total=Sum('cari', default=0))['total'] or 0
    gider_mehmet_havale = qs.filter(hareket_tipi='gider').aggregate(total=Sum('mehmet_havale', default=0))['total'] or 0
    
    # Net toplamlar (gelir - gider)
    nakit_net = float(gelir_nakit - gider_nakit)
    kredi_karti_net = float(gelir_kredi_karti - gider_kredi_karti)
    cari_net = float(gelir_cari - gider_cari)
    mehmet_havale_field_net = float(gelir_mehmet_havale - gider_mehmet_havale)
    
    # Nakit → Çanta'ya ekle
    canta_toplam += nakit_net
    
    # Kredi Kartı ve Cari için kasa adına göre dağıt (zaten çalışıyor)
    # Mehmet Havale için hem kasa hem field toplamını kontrol et
    if mehmet_havale_toplam == 0:
        mehmet_havale_toplam = mehmet_havale_field_net
    else:
        # Hem kasa hem field varsa, field toplamını ekle
        mehmet_havale_toplam += mehmet_havale_field_net
    
    # Genel toplamı güncelle (tüm kasaların toplamı)
    genel_toplam = sum([servis_toplam, merkez_satis_toplam, canta_toplam, mehmet_havale_toplam])

    context = {
        'page_title': 'Gelir/Gider Toplamları',
        'servis_toplam': servis_toplam,
        'merkez_satis_toplam': merkez_satis_toplam,
        'canta_toplam': canta_toplam,
        'mehmet_havale_toplam': mehmet_havale_toplam,
        'genel_toplam': genel_toplam,
        'kasa_satirlari': rows,
        # Ödeme yöntemleri toplamları (debug/ekstra bilgi için)
        'odeme_toplamlari': {
            'nakit': nakit_net,
            'kredi_karti': kredi_karti_net,
            'cari': cari_net,
            'mehmet_havale': mehmet_havale_field_net,
        },
    }
    return render(request, 'dashboard/messages.html', context)

@login_required
def finance_overview(request):
    """JSON: Son 12 ay için aylık gelir/gider toplamları (canlı)."""
    try:
        labels = []
        income = []
        expense = []
        toplam_ifade = (F('nakit') + F('kredi_karti') + F('cari') + F('mehmet_havale'))

        now = timezone.now()
        for i in range(11, -1, -1):
            start_date = (now - timedelta(days=30*i)).date().replace(day=1)
            # Bir sonraki ayın 1'i
            if start_date.month == 12:
                next_month = start_date.replace(year=start_date.year + 1, month=1, day=1)
            else:
                next_month = start_date.replace(month=start_date.month + 1, day=1)

            gelir = Transaction.objects.filter(
                created_by=request.user,
                tarih__gte=start_date,
                tarih__lt=next_month,
                hareket_tipi='gelir'
            ).aggregate(total=Sum(toplam_ifade))['total'] or 0

            gider = Transaction.objects.filter(
                created_by=request.user,
                tarih__gte=start_date,
                tarih__lt=next_month,
                hareket_tipi='gider'
            ).aggregate(total=Sum(toplam_ifade))['total'] or 0

            labels.append(start_date.strftime('%b'))
            income.append(float(gelir))
            expense.append(float(gider))

        return JsonResponse({
            'success': True,
            'labels': labels,
            'income': income,
            'expense': expense,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

def calendar(request):
    """Calendar sayfası"""
    context = {
        'page_title': 'Takvim',
        'events': [],
    }
    return render(request, 'dashboard/calendar.html', context)

@login_required
def create_event(request):
    """Etkinlik oluşturma API endpoint'i"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Gerekli alanları kontrol et
            required_fields = ['title', 'date', 'time']
            for field in required_fields:
                if not data.get(field):
                    return JsonResponse({
                        'success': False,
                        'error': f'{field} alanı gereklidir'
                    }, status=400)
            
            # Tarih ve saat verilerini parse et
            from datetime import datetime
            date_obj = datetime.strptime(data['date'], '%Y-%m-%d').date()
            time_obj = datetime.strptime(data['time'], '%H:%M').time()
            
            # Event oluştur ve veritabanına kaydet
            event = Event.objects.create(
                title=data.get('title', ''),
                description=data.get('description', ''),
                type=data.get('type', 'event'),
                priority=data.get('priority', 'medium'),
                date=date_obj,
                time=time_obj,
                duration=int(data.get('duration', 60)),
                location=data.get('location', ''),
                attendees=data.get('attendees', ''),
                recurring=data.get('recurring', False),
                recurrence=data.get('recurrence', 'none'),
                reminders=json.dumps(data.get('reminders', ['15'])),
                created_by=request.user
            )
            
            # Bildirim oluştur
            try:
                create_event_notifications(event)
            except Exception as e:
                print(f"Bildirim oluşturma hatası: {e}")
            
            return JsonResponse({
                'success': True,
                'message': 'Etkinlik başarıyla oluşturuldu!',
                'event': {
                    'id': event.id,
                    'title': event.title,
                    'date': event.date.strftime('%Y-%m-%d'),
                    'time': event.time.strftime('%H:%M'),
                    'description': event.description,
                    'location': event.location,
                    'type': event.type,
                    'priority': event.priority
                }
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Geçersiz JSON verisi'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'error': 'Sadece POST istekleri kabul edilir'
    }, status=405)

@login_required
def get_events(request):
    """Etkinlikleri getiren API endpoint'i"""
    try:
        # Tarih filtreleri
        start_date = request.GET.get('start')
        end_date = request.GET.get('end')
        
        events = Event.objects.filter(created_by=request.user)
        
        if start_date:
            events = events.filter(date__gte=start_date)
        if end_date:
            events = events.filter(date__lte=end_date)
        
        events_data = []
        for event in events:
            events_data.append({
                'id': event.id,
                'title': event.title,
                'type': event.type,
                'date': event.date.strftime('%Y-%m-%d'),
                'time': event.time.strftime('%H:%M'),
                'timeStr': event.time.strftime('%H:%M'),
                'dateStr': event.date.strftime('%d %b'),
                'description': event.description,
                'location': event.location,
                'priority': event.priority,
                'duration': event.duration,
                'attendees': event.attendees,
                'recurring': event.recurring,
                'recurrence': event.recurrence,
                'dateObj': event.date.isoformat(),
                'read': True
            })
        
        return JsonResponse({
            'success': True,
            'events': events_data
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def kategoriler(request):
    """Kategoriler sayfası - Kategori yönetimi"""
    if request.method == 'POST':
        # Yeni kategori ekleme
        kategori_adi = request.POST.get('kategori_adi')
        parent_id = request.POST.get('parent_id')
        
        if kategori_adi:
            parent = None
            if parent_id:
                try:
                    parent = TransactionCategory.objects.get(id=parent_id)
                except TransactionCategory.DoesNotExist:
                    pass
            
            TransactionCategory.objects.create(
                name=kategori_adi,
                parent=parent,
                created_by=request.user
            )
            if parent:
                messages.success(request, f'Alt kategori "{kategori_adi}" başarıyla "{parent.name}" kategorisinin altına eklendi!')
            else:
                messages.success(request, f'Ana kategori "{kategori_adi}" başarıyla eklendi!')
        else:
            messages.error(request, 'Kategori adı boş olamaz!')
        
        return redirect('dashboard:kategoriler')
    
    # Kategorileri hiyerarşik olarak getir
    ana_kategoriler = TransactionCategory.objects.filter(parent=None).order_by('name')
    alt_kategoriler = TransactionCategory.objects.filter(parent__isnull=False).order_by('name')
    
    # Hiyerarşik sıralama: Ana kategori -> Alt kategorileri -> Sonraki ana kategori
    tum_kategoriler = []
    for ana_kategori in ana_kategoriler:
        # Ana kategoriyi ekle
        tum_kategoriler.append(ana_kategori)
        # Bu ana kategorinin alt kategorilerini ekle
        alt_kategoriler_bu_ana = TransactionCategory.objects.filter(parent=ana_kategori).order_by('name')
        tum_kategoriler.extend(alt_kategoriler_bu_ana)
    
    # Tüm kullanıcıları getir (dropdown için)
    tum_kullanicilar = User.objects.all().order_by('username')
    
    context = {
        'page_title': 'Kategoriler',
        'ana_kategoriler': ana_kategoriler,
        'tum_kategoriler': tum_kategoriler,
        'alt_kategoriler': alt_kategoriler,
        'tum_kullanicilar': tum_kullanicilar,
        'stats': {
            'toplam': len(tum_kategoriler),
            'ana': ana_kategoriler.count(),
            'alt': alt_kategoriler.count(),
            'aktif': len(tum_kategoriler),
        }
    }
    return render(request, 'dashboard/kategoriler.html', context)

def settings(request):
    """Settings sayfası"""
    context = {
        'page_title': 'Settings',
    }
    return render(request, 'dashboard/settings.html', context)

def security(request):
    """Security sayfası"""
    context = {
        'page_title': 'Security',
    }
    return render(request, 'dashboard/security.html', context)

def help(request):
    """Help sayfası"""
    context = {
        'page_title': 'Help & Support',
    }
    return render(request, 'dashboard/help.html', context)

def settings(request):
    """Settings sayfası"""
    context = {
        'page_title': 'Settings',
    }
    return render(request, 'dashboard/settings.html', context)

def security(request):
    """Security sayfası"""
    context = {
        'page_title': 'Security',
    }
    return render(request, 'dashboard/security.html', context)

def help(request):
    """Help sayfası"""
    context = {
        'page_title': 'Help & Support',
    }
    return render(request, 'dashboard/help.html', context)

@login_required
def export_excel(request):
    """Sipariş Envanterini Excel'e aktar"""
    # Filtreleme parametrelerini al
    firma = request.GET.get('firma', '')
    marka = request.GET.get('marka', '')
    grup = request.GET.get('grup', '')
    durum = request.GET.get('durum', '')
    mevsim = request.GET.get('mevsim', '')
    ambar = request.GET.get('ambar', '')
    tarih_filtre = request.GET.get('tarih', '')
    baslangic_tarihi = request.GET.get('baslangic_tarihi', '')
    bitis_tarihi = request.GET.get('bitis_tarihi', '')
    
    # Siparişleri filtrele (iptal edilenleri ve kontrol edilenleri hariç tut, sadece kullanıcının siparişleri)
    siparisler = Siparis.objects.filter(user=request.user).exclude(durum__in=['iptal', 'kontrol'])
    
    if firma:
        siparisler = siparisler.filter(cari_firma__icontains=firma)
    if marka:
        siparisler = siparisler.filter(marka__icontains=marka)
    if grup:
        siparisler = siparisler.filter(grup=grup)
    if durum:
        siparisler = siparisler.filter(durum=durum)
    if mevsim:
        siparisler = siparisler.filter(mevsim=mevsim)
    if ambar:
        siparisler = siparisler.filter(ambar=ambar)
    
    # Tarih filtreleme uygula
    now = timezone.now()
    if tarih_filtre:
        if tarih_filtre == 'son-1-ay':
            start_date = now - timedelta(days=30)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'son-3-ay':
            start_date = now - timedelta(days=90)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'son-6-ay':
            start_date = now - timedelta(days=180)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'bugun':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = now.replace(hour=23, minute=59, second=59, microsecond=999999)
            siparisler = siparisler.filter(olusturma_tarihi__range=[start_date, end_date])
        elif tarih_filtre == 'bu-hafta':
            start_date = now - timedelta(days=now.weekday())
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'bu-ay':
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
    
    # Özel tarih aralığı filtreleme
    if baslangic_tarihi and bitis_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0))
            end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
            siparisler = siparisler.filter(olusturma_tarihi__range=[start_date, end_date])
        except ValueError:
            pass
    elif baslangic_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0))
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        except ValueError:
            pass
    elif bitis_tarihi:
        try:
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d')
            end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
            siparisler = siparisler.filter(olusturma_tarihi__lte=end_date)
        except ValueError:
            pass
    
    # Excel dosyası oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Sipariş Envanteri"
    
    # Başlık satırı
    headers = [
        'CARI (FIRMA)', 'ÜRÜN', 'MARKA', 'GRUP', 'MEVSİM', 'ADET', 
        'BİRİM FİYAT', 'DURUM', 'AMBAR', 'AÇIKLAMA 1', 'TOPLAM FİYAT', 
        'ÖDEME', 'SMS', 'ÖNE ÇIKAR', 'OLUŞTURMA TARİHİ', 'GÜNCELLEME TARİHİ'
    ]
    
    # Başlık stilini ayarla
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Veri satırları
    for row, siparis in enumerate(siparisler, 2):
        ws.cell(row=row, column=1, value=siparis.cari_firma)
        ws.cell(row=row, column=2, value=siparis.urun)
        ws.cell(row=row, column=3, value=siparis.marka)
        ws.cell(row=row, column=4, value=siparis.get_grup_display())
        ws.cell(row=row, column=5, value=siparis.get_mevsim_display())
        ws.cell(row=row, column=6, value=siparis.adet)
        ws.cell(row=row, column=7, value=siparis.birim_fiyat)
        ws.cell(row=row, column=8, value=siparis.get_durum_display())
        ws.cell(row=row, column=9, value=siparis.get_ambar_display())
        ws.cell(row=row, column=10, value=siparis.aciklama)
        ws.cell(row=row, column=11, value=siparis.toplam_fiyat)
        ws.cell(row=row, column=12, value=siparis.get_odeme_display())
        ws.cell(row=row, column=13, value=siparis.get_sms_durum_display())
        ws.cell(row=row, column=14, value="Evet" if siparis.one_cikar else "Hayır")
        ws.cell(row=row, column=15, value=siparis.olusturma_tarihi.strftime('%d.%m.%Y %H:%M'))
        ws.cell(row=row, column=16, value=siparis.guncelleme_tarihi.strftime('%d.%m.%Y %H:%M'))
    
    # Sütun genişliklerini ayarla
    column_widths = [20, 25, 15, 12, 12, 8, 12, 15, 10, 30, 12, 12, 10, 12, 18, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # HTTP response oluştur
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="siparis_envanteri.xlsx"'
    
    # Excel dosyasını response'a yaz
    wb.save(response)
    return response

@login_required
def export_cancelled_excel(request):
    """İptal Edilen Siparişleri Excel'e aktar"""
    # Filtreleme parametrelerini al
    firma = request.GET.get('firma', '')
    marka = request.GET.get('marka', '')
    grup = request.GET.get('grup', '')
    mevsim = request.GET.get('mevsim', '')
    ambar = request.GET.get('ambar', '')
    tarih_filtre = request.GET.get('tarih', '')
    baslangic_tarihi = request.GET.get('baslangic_tarihi', '')
    bitis_tarihi = request.GET.get('bitis_tarihi', '')
    
    # Sadece iptal edilen siparişleri filtrele (sadece kullanıcının siparişleri)
    siparisler = Siparis.objects.filter(durum='iptal', user=request.user)
    
    if firma:
        siparisler = siparisler.filter(cari_firma__icontains=firma)
    if marka:
        siparisler = siparisler.filter(marka__icontains=marka)
    if grup:
        siparisler = siparisler.filter(grup=grup)
    if mevsim:
        siparisler = siparisler.filter(mevsim=mevsim)
    if ambar:
        siparisler = siparisler.filter(ambar=ambar)
    
    # Tarih filtreleme uygula
    now = timezone.now()
    if tarih_filtre:
        if tarih_filtre == 'son-1-ay':
            start_date = now - timedelta(days=30)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'son-3-ay':
            start_date = now - timedelta(days=90)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'son-6-ay':
            start_date = now - timedelta(days=180)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'bugun':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = now.replace(hour=23, minute=59, second=59, microsecond=999999)
            siparisler = siparisler.filter(olusturma_tarihi__range=[start_date, end_date])
        elif tarih_filtre == 'bu-hafta':
            start_date = now - timedelta(days=now.weekday())
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'bu-ay':
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
    
    # Özel tarih aralığı filtreleme
    if baslangic_tarihi and bitis_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0))
            end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
            siparisler = siparisler.filter(olusturma_tarihi__range=[start_date, end_date])
        except ValueError:
            pass
    elif baslangic_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0))
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        except ValueError:
            pass
    elif bitis_tarihi:
        try:
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d')
            end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
            siparisler = siparisler.filter(olusturma_tarihi__lte=end_date)
        except ValueError:
            pass
    
    # Excel dosyası oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "İptal Edilen Siparişler"
    
    # Başlık satırı
    headers = [
        'ID', 'CARI (FIRMA)', 'ÜRÜN', 'MARKA', 'GRUP', 'MEVSİM', 'ADET', 
        'BİRİM FİYAT', 'AMBAR', 'AÇIKLAMA', 'TOPLAM FİYAT', 
        'ÖDEME', 'SMS', 'ÖNE ÇIKAR', 'OLUŞTURMA TARİHİ', 'İPTAL TARİHİ'
    ]
    
    # Başlık stilini ayarla
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")  # Kırmızı renk
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Veri satırları
    for row, siparis in enumerate(siparisler, 2):
        ws.cell(row=row, column=1, value=siparis.id)
        ws.cell(row=row, column=2, value=siparis.cari_firma)
        ws.cell(row=row, column=3, value=siparis.urun)
        ws.cell(row=row, column=4, value=siparis.marka)
        ws.cell(row=row, column=5, value=siparis.get_grup_display())
        ws.cell(row=row, column=6, value=siparis.get_mevsim_display())
        ws.cell(row=row, column=7, value=siparis.adet)
        ws.cell(row=row, column=8, value=siparis.birim_fiyat)
        ws.cell(row=row, column=9, value=siparis.get_ambar_display())
        ws.cell(row=row, column=10, value=siparis.aciklama)
        ws.cell(row=row, column=11, value=siparis.toplam_fiyat)
        ws.cell(row=row, column=12, value=siparis.get_odeme_display())
        ws.cell(row=row, column=13, value=siparis.get_sms_durum_display())
        ws.cell(row=row, column=14, value="Evet" if siparis.one_cikar else "Hayır")
        ws.cell(row=row, column=15, value=siparis.olusturma_tarihi.strftime('%d.%m.%Y %H:%M'))
        ws.cell(row=row, column=16, value=siparis.guncelleme_tarihi.strftime('%d.%m.%Y %H:%M'))
    
    # Sütun genişliklerini ayarla
    column_widths = [8, 20, 25, 15, 12, 12, 8, 12, 10, 30, 12, 12, 10, 12, 18, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # HTTP response oluştur
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="iptal_edilen_siparisler.xlsx"'
    
    # Excel dosyasını response'a yaz
    wb.save(response)
    return response

@login_required
def export_checked_excel(request):
    """Kontrol Edilen Siparişleri Excel'e aktar"""
    # Filtreleme parametrelerini al
    firma = request.GET.get('firma', '')
    marka = request.GET.get('marka', '')
    grup = request.GET.get('grup', '')
    mevsim = request.GET.get('mevsim', '')
    ambar = request.GET.get('ambar', '')
    tarih_filtre = request.GET.get('tarih', '')
    baslangic_tarihi = request.GET.get('baslangic_tarihi', '')
    bitis_tarihi = request.GET.get('bitis_tarihi', '')
    
    # Sadece kontrol edilen siparişleri filtrele (sadece kullanıcının siparişleri)
    siparisler = Siparis.objects.filter(durum='kontrol', user=request.user)
    
    if firma:
        siparisler = siparisler.filter(cari_firma__icontains=firma)
    if marka:
        siparisler = siparisler.filter(marka__icontains=marka)
    if grup:
        siparisler = siparisler.filter(grup=grup)
    if mevsim:
        siparisler = siparisler.filter(mevsim=mevsim)
    if ambar:
        siparisler = siparisler.filter(ambar=ambar)
    
    # Tarih filtreleme uygula
    now = timezone.now()
    if tarih_filtre:
        if tarih_filtre == 'son-1-ay':
            start_date = now - timedelta(days=30)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'son-3-ay':
            start_date = now - timedelta(days=90)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'son-6-ay':
            start_date = now - timedelta(days=180)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'bugun':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = now.replace(hour=23, minute=59, second=59, microsecond=999999)
            siparisler = siparisler.filter(olusturma_tarihi__range=[start_date, end_date])
        elif tarih_filtre == 'bu-hafta':
            start_date = now - timedelta(days=now.weekday())
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'bu-ay':
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
    
    # Özel tarih aralığı filtreleme
    if baslangic_tarihi and bitis_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0))
            end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
            siparisler = siparisler.filter(olusturma_tarihi__range=[start_date, end_date])
        except ValueError:
            pass
    elif baslangic_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0))
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        except ValueError:
            pass
    elif bitis_tarihi:
        try:
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d')
            end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
            siparisler = siparisler.filter(olusturma_tarihi__lte=end_date)
        except ValueError:
            pass
    
    # Excel dosyası oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Kontrol Edilen Siparişler"
    
    # Başlık satırı
    headers = [
        'CARI (FIRMA)', 'ÜRÜN', 'MARKA', 'GRUP', 'MEVSİM', 'ADET', 
        'BİRİM FİYAT', 'AMBAR', 'AÇIKLAMA', 'TOPLAM FİYAT', 
        'ÖDEME', 'SMS', 'ÖNE ÇIKAR', 'OLUŞTURMA TARİHİ', 'KONTROL TARİHİ'
    ]
    
    # Başlık stilini ayarla
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")  # Sarı renk
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Veri satırları
    for row, siparis in enumerate(siparisler, 2):
        ws.cell(row=row, column=1, value=siparis.cari_firma)
        ws.cell(row=row, column=2, value=siparis.urun)
        ws.cell(row=row, column=3, value=siparis.marka)
        ws.cell(row=row, column=4, value=siparis.get_grup_display())
        ws.cell(row=row, column=5, value=siparis.get_mevsim_display())
        ws.cell(row=row, column=6, value=siparis.adet)
        ws.cell(row=row, column=7, value=siparis.birim_fiyat)
        ws.cell(row=row, column=8, value=siparis.get_ambar_display())
        ws.cell(row=row, column=9, value=siparis.aciklama)
        ws.cell(row=row, column=10, value=siparis.toplam_fiyat)
        ws.cell(row=row, column=11, value=siparis.get_odeme_display())
        ws.cell(row=row, column=12, value=siparis.get_sms_durum_display())
        ws.cell(row=row, column=13, value="Evet" if siparis.one_cikar else "Hayır")
        ws.cell(row=row, column=14, value=siparis.olusturma_tarihi.strftime('%d.%m.%Y %H:%M'))
        ws.cell(row=row, column=15, value=siparis.guncelleme_tarihi.strftime('%d.%m.%Y %H:%M'))
    
    # Sütun genişliklerini ayarla
    column_widths = [20, 25, 15, 12, 12, 8, 12, 10, 30, 12, 12, 10, 12, 18, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # HTTP response oluştur
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="kontrol_edilen_siparisler.xlsx"'
    
    # Excel dosyasını response'a yaz
    wb.save(response)
    return response

def login_view(request):
    """Kullanıcı giriş sayfası"""
    if request.user.is_authenticated:
        return redirect('dashboard:index')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        if username and password:
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f'Hoş geldiniz, {user.first_name or user.username}!')
                next_url = request.GET.get('next', 'dashboard:index')
                return redirect(next_url)
            else:
                messages.error(request, 'Kullanıcı adı veya şifre hatalı!')
        else:
            messages.error(request, 'Lütfen tüm alanları doldurun!')
    
    context = {
        'page_title': 'Giriş Yap',
    }
    return render(request, 'dashboard/login.html', context)

def logout_view(request):
    """Kullanıcı çıkış"""
    if request.user.is_authenticated:
        username = request.user.first_name or request.user.username
        logout(request)
        messages.success(request, f'Güle güle, {username}!')
    return redirect('dashboard:login')

# Bildirim API'leri

@login_required
def get_notifications(request):
    """Kullanıcının bildirimlerini getiren API endpoint'i"""
    try:
        # Kullanıcının bildirimlerini getir
        notifications = Notification.objects.filter(user=request.user).order_by('-scheduled_time')[:20]
        
        notifications_data = []
        for notification in notifications:
            notifications_data.append({
                'id': notification.id,
                'title': notification.title,
                'message': notification.message,
                'type': notification.type,
                'status': notification.status,
                'icon': notification.get_type_icon(),
                'color': notification.get_type_color(),
                'scheduled_time': notification.scheduled_time.isoformat(),
                'sent_time': notification.sent_time.isoformat() if notification.sent_time else None,
                'read_time': notification.read_time.isoformat() if notification.read_time else None,
                'event_id': None,  # Event geçici olarak devre dışı
                'is_overdue': notification.is_overdue(),
                'extra_data': notification.get_extra_data_dict()
            })
        
        # Okunmamış bildirim sayısı
        unread_count = Notification.objects.filter(
            user=request.user, 
            status__in=['pending', 'sent']
        ).count()
        
        return JsonResponse({
            'success': True,
            'notifications': notifications_data,
            'unread_count': unread_count
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def mark_notification_read(request, notification_id):
    """Bildirimi okundu olarak işaretle"""
    try:
        notification = get_object_or_404(Notification, id=notification_id, user=request.user)
        notification.mark_as_read()
        
        return JsonResponse({
            'success': True,
            'message': 'Bildirim okundu olarak işaretlendi'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def mark_all_notifications_read(request):
    """Tüm bildirimleri okundu olarak işaretle"""
    try:
        notifications = Notification.objects.filter(
            user=request.user, 
            status__in=['pending', 'sent']
        )
        
        for notification in notifications:
            notification.mark_as_read()
        
        return JsonResponse({
            'success': True,
            'message': f'{notifications.count()} bildirim okundu olarak işaretlendi'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def dismiss_notification(request, notification_id):
    """Bildirimi kapat"""
    try:
        notification = get_object_or_404(Notification, id=notification_id, user=request.user)
        notification.status = 'dismissed'
        notification.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Bildirim kapatıldı'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

def create_event_notifications(event):
    """Etkinlik için bildirimler oluştur"""
    from datetime import datetime, timedelta
    from django.utils import timezone
    
    try:
        # Hatırlatıcı zamanlarını parse et
        reminders = json.loads(event.reminders) if event.reminders else ['15']
        
        # Etkinlik tarih ve saatini birleştir
        naive_datetime = datetime.combine(event.date, event.time)
        
        # Timezone aware yap
        if timezone.is_naive(naive_datetime):
            event_datetime = timezone.make_aware(naive_datetime)
        else:
            event_datetime = naive_datetime
        
        now = timezone.now()
        
        # Her hatırlatıcı için bildirim oluştur
        for reminder_minutes in reminders:
            try:
                minutes = int(reminder_minutes)
                reminder_time = event_datetime - timedelta(minutes=minutes)
                
                # Geçmiş tarih kontrolü
                if reminder_time > now:
                    # Hatırlatıcı bildirimi oluştur
                    Notification.objects.create(
                        title=f"Etkinlik Hatırlatıcısı: {event.title}",
                        message=f"{event.title} etkinliği {minutes} dakika sonra başlayacak. Konum: {event.location or 'Belirtilmemiş'}",
                        type='event_reminder',
                        user=event.created_by,
                        scheduled_time=reminder_time,
                        extra_data=json.dumps({
                            'event_id': event.id,
                            'reminder_minutes': minutes,
                            'event_type': event.type,
                            'event_priority': event.priority
                        })
                    )
            except (ValueError, TypeError):
                continue
        
        # Etkinlik başlangıç bildirimi
        if event_datetime > now:
            Notification.objects.create(
                title=f"Etkinlik Başlıyor: {event.title}",
                message=f"{event.title} etkinliği şimdi başlıyor! Süre: {event.get_duration_display()}",
                type='event_start',
                user=event.created_by,
                scheduled_time=event_datetime,
                extra_data=json.dumps({
                    'event_id': event.id,
                    'event_type': event.type,
                    'event_priority': event.priority,
                    'duration': event.duration
                })
            )
        
        # Hemen bir bildirim de oluştur (etkinlik oluşturuldu)
        Notification.objects.create(
            title=f"Yeni Etkinlik Oluşturuldu",
            message=f"'{event.title}' etkinliği {event.date.strftime('%d.%m.%Y')} tarihinde {event.time.strftime('%H:%M')} saatinde oluşturuldu.",
            type='event_created',
            user=event.created_by,
            status='sent',
            extra_data=json.dumps({
                'event_id': event.id,
                'event_type': event.type,
                'event_priority': event.priority
            })
        )
            
    except Exception as e:
        # Hata logla
        print(f"Bildirim oluşturma hatası: {e}")
        pass

# Etkinlik oluşturma view'ını güncelle (devre dışı sürüm kullanılmıyor)
# Aşağıdaki sürüm kaldırıldı; aktif olan basit `create_event` yukarıda tanımlıdır.

@login_required
def finance(request):
    """Gelir/Gider İşlemleri sayfası"""
    if request.method == 'POST':
        form = TransactionForm(request.POST)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.created_by = request.user
            transaction.save()
            
            messages.success(request, 'İşlem başarıyla kaydedildi!')
            
            # next parametresi varsa oraya yönlendir
            next_url = request.POST.get('next')
            if next_url:
                return redirect(next_url)
            else:
                return redirect('dashboard:finance')
        else:
            messages.error(request, 'Form hatası! Lütfen alanları kontrol edin.')
    else:
        form = TransactionForm()
    
    context = {
        'page_title': 'Gelir/Gider İşlemleri',
        'form': form,
    }
    return render(request, 'dashboard/finance.html', context)

@login_required
def kategori_sil(request, kategori_id):
    """Kategori silme"""
    try:
        kategori = get_object_or_404(TransactionCategory, id=kategori_id)
        kategori_adi = kategori.name
        kategori.delete()
        messages.success(request, f'Kategori "{kategori_adi}" başarıyla silindi!')
    except Exception as e:
        messages.error(request, f'Kategori silinirken hata oluştu: {str(e)}')
    
    return redirect('dashboard:kategoriler')

@login_required
def kategori_duzenle(request, kategori_id):
    """Kategori düzenleme"""
    kategori = get_object_or_404(TransactionCategory, id=kategori_id)
    
    if request.method == 'POST':
        kategori_adi = request.POST.get('kategori_adi')
        parent_id = request.POST.get('parent_id')
        kaydi_acan_id = request.POST.get('kaydi_acan_id')
        
        if kategori_adi:
            # Üst kategori kontrolü
            parent = None
            if parent_id:
                try:
                    parent = TransactionCategory.objects.get(id=parent_id)
                    # Kendi kendisinin alt kategorisi olamaz
                    if parent.id == kategori.id:
                        messages.error(request, 'Kategori kendi kendisinin alt kategorisi olamaz!')
                        return redirect('dashboard:kategoriler')
                except TransactionCategory.DoesNotExist:
                    pass
            
            # Kaydı açan kullanıcıyı güncelle
            if kaydi_acan_id:
                try:
                    user = User.objects.get(id=kaydi_acan_id)
                    kategori.created_by = user
                except User.DoesNotExist:
                    messages.warning(request, 'Seçilen kullanıcı bulunamadı. Kaydı açan değiştirilmedi.')
            
            # Kategoriyi güncelle
            eski_ad = kategori.name
            kategori.name = kategori_adi
            kategori.parent = parent
            kategori.save()
            
            messages.success(request, f'Kategori "{eski_ad}" → "{kategori_adi}" olarak güncellendi!')
        else:
            messages.error(request, 'Kategori adı boş olamaz!')
    
    return redirect('dashboard:kategoriler')

@login_required
def get_alt_kategoriler(request, ana_kategori_id):
    """Ana kategoriye ait alt kategorileri JSON olarak döndür"""
    try:
        ana_kategori = get_object_or_404(TransactionCategory, id=ana_kategori_id, parent=None)
        alt_kategoriler = ana_kategori.children.all().order_by('name')
        
        data = []
        for alt_kategori in alt_kategoriler:
            data.append({
                'id': alt_kategori.id,
                'name': alt_kategori.name
            })
        
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def brand_distribution_api(request):
    """Lastik marka dağılımı API endpoint'i - Canlı veri"""
    try:
        # Kullanıcının siparişlerinden marka dağılımını hesapla
        brand_distribution = Siparis.objects.filter(
            user=request.user
        ).values('marka').annotate(
            total_adet=Sum('adet')
        ).order_by('-total_adet')
        
        # Veri hazırla
        brands = []
        colors = [
            '#3b82f6', '#ef4444', '#10b981', '#f59e0b', 
            '#8b5cf6', '#ec4899', '#06b6d4', '#84cc16',
            '#f97316', '#06b6d4', '#8b5cf6', '#ef4444'
        ]
        
        for i, brand in enumerate(brand_distribution):
            if brand['marka'] and brand['total_adet']:  # Boş marka adlarını filtrele
                brands.append({
                    'name': brand['marka'],
                    'count': brand['total_adet'],
                    'color': colors[i % len(colors)]
                })
        
        return JsonResponse({
            'success': True,
            'brands': brands,
            'total_brands': len(brands),
            'total_count': sum(brand['count'] for brand in brands)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def malzeme_excel_upload(request):
    if request.method == 'POST':
        form = MalzemeExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']
            try:
                from openpyxl import load_workbook
                wb = load_workbook(excel_file)
                ws = wb.active
                
                # Create MalzemeDosya first
                dosya = MalzemeDosya.objects.create(
                    dosya_adi=excel_file.name,
                    kullanici=request.user
                )
                
                # Get headers from first row
                headers = []
                for cell in ws[1]:
                    if cell.value:
                        headers.append(str(cell.value).strip())
                
                eklenen = 0
                for row_num in range(2, ws.max_row + 1):
                    try:
                        # Create row dict
                        row_data = {}
                        for col_num, header in enumerate(headers, 1):
                            if col_num <= len(headers):
                                cell_value = ws.cell(row=row_num, column=col_num).value
                                row_data[header] = cell_value
                        
                        tarih = row_data.get('TARİH') or row_data.get('TARIH') or ''
                        # Handle different date formats
                        if isinstance(tarih, datetime.datetime):
                            tarih = tarih.date()
                        elif isinstance(tarih, str) and tarih and '.' in tarih:
                            try:
                                tarih = datetime.datetime.strptime(tarih, '%d.%m.%Y').date()
                            except:
                                tarih = date.today()
                        elif isinstance(tarih, str) and tarih and '-' in tarih:
                            try:
                                tarih = datetime.datetime.strptime(tarih, '%Y-%m-%d').date()
                            except:
                                tarih = date.today()
                        else:
                            tarih = date.today()

                        tutar_raw = row_data.get('TUTAR') or '0'
                        try:
                            tutar = float(str(tutar_raw).replace('.', '').replace(',', '.')) if tutar_raw else 0
                        except:
                            tutar = 0
                            
                        hareket = MalzemeHareketi(
                            dosya=dosya,
                            tarih=tarih,
                            faturano=str(row_data.get('FATURA NO') or row_data.get('FATURANO') or '')[:100],
                            musteri=str(row_data.get('MÜŞTERİ') or row_data.get('MÜŞTERI') or '')[:255],
                            urun=str(row_data.get('ÜRÜN') or row_data.get('URUN') or '')[:255],
                            tutar=tutar,
                            odeme_sekli=str(row_data.get('ÖDEME ŞEKLİ') or row_data.get('ÖDEME PLANI') or '')[:100],
                            kullanici=request.user,
                        )
                        hareket.save()
                        eklenen += 1
                    except Exception as e:
                        print(f"Row error: {e}")
                        continue
                        
                messages.success(request, f"Başarıyla {eklenen} satır kaydedildi!")
                return redirect('dashboard:products')
            except Exception as e:
                print(f"Excel upload error: {e}")
                messages.error(request, f"Excel yükleme hatası: {str(e)}")
                return redirect('dashboard:finance')
                messages.error(request, f"Dosya okunamadı: {str(e)}")
        else:
            messages.error(request, "Lütfen geçerli bir dosya seçin.")
    else:
        form = MalzemeExcelUploadForm()
    return render(request, 'dashboard/malzeme_excel_upload.html', {'form': form})

@csrf_exempt
@login_required
def malzeme_excel_kaydet(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Sadece POST!'}, status=400)
    try:
        # JSON parsing
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError as e:
            return JsonResponse({'success': False, 'error': f'JSON parsing hatası: {str(e)}'}, status=400)
        
        filename = data.get('filename')
        rows = data.get('rows', [])
        
        # Debug: Gelen veriyi kontrol et
        print(f"Gelen dosya: {filename}")
        print(f"Gelen satır sayısı: {len(rows)}")
        print(f"Request body boyutu: {len(request.body)} bytes")
        if rows:
            print(f"İlk satır örneği: {rows[0]}")
            print(f"İlk satır keys: {list(rows[0].keys()) if isinstance(rows[0], dict) else 'Not dict'}")
        
        if not filename:
            return JsonResponse({'success': False, 'error': 'Dosya adı eksik.'}, status=400)
        if not isinstance(rows, list):
            return JsonResponse({'success': False, 'error': f'Rows veri tipi yanlış: {type(rows)}'}, status=400)
        if len(rows) == 0:
            return JsonResponse({'success': False, 'error': 'Excel dosyasında veri bulunamadı.'}, status=400)
            
        dosya = MalzemeDosya.objects.create(dosya_adi=filename, kullanici=request.user)
        print(f"Dosya oluşturuldu: ID={dosya.id}")
        
        eklenen = 0
        hatalar = []
        
        for i, row in enumerate(rows):
            try:
                # Debug: Satır verilerini kontrol et
                print(f"Satır {i+1}: {row}")
                
                if not row or not isinstance(row, dict):
                    print(f"Satır {i+1} geçersiz: {row}")
                    continue
                
                # Tarih işleme - daha esnek
                tarih_raw = None
                for key in row.keys():
                    if any(x in key.upper() for x in ['TARİH', 'TARIH', 'DATE']):
                        tarih_raw = row[key]
                        break
                
                tarih = date.today()  # Default tarih
                
                if tarih_raw:
                    try:
                        if isinstance(tarih_raw, str) and tarih_raw.strip():
                            tarih_str = tarih_raw.strip()
                            if '.' in tarih_str:
                                # DD.MM.YYYY veya DD.MM.YY formatı
                                parts = tarih_str.split('.')
                                if len(parts) == 3:
                                    day, month, year = parts
                                    if len(year) == 2:
                                        year = '20' + year if int(year) < 50 else '19' + year
                                    tarih = date(int(year), int(month), int(day))
                            elif '-' in tarih_str:
                                # YYYY-MM-DD formatı
                                tarih = datetime.datetime.strptime(tarih_str, '%Y-%m-%d').date()
                        elif isinstance(tarih_raw, (int, float)) and tarih_raw > 0:
                            # Excel serial date
                            tarih = datetime.fromordinal(date(1900,1,1).toordinal() + int(tarih_raw) - 2).date()
                    except Exception as e:
                        print(f"Tarih parse hatası: {e}, raw: {tarih_raw}")
                        tarih = date.today()
                
                # Tüm sütun adlarını kontrol et ve esnek eşleştirme yap
                row_keys = list(row.keys())
                print(f"Satır {i+1} sütunları: {row_keys}")
                
                # Esnek sütun eşleştirmesi - daha geniş arama
                faturano = ''
                for key in row_keys:
                    key_upper = key.upper().strip()
                    if any(x in key_upper for x in ['FATURA', 'INVOICE', 'NO', 'BELGE']):
                        val = row[key]
                        faturano = str(val).strip() if val is not None else ''
                        break
                
                musteri = ''
                for key in row_keys:
                    key_upper = key.upper().strip()
                    if any(x in key_upper for x in ['MÜŞTERİ', 'MÜŞTERI', 'MUSTERI', 'CUSTOMER', 'CLIENT', 'CARİ', 'CARI']):
                        val = row[key]
                        musteri = str(val).strip() if val is not None else ''
                        break
                
                urun = ''
                for key in row_keys:
                    key_upper = key.upper().strip()
                    if any(x in key_upper for x in ['ÜRÜN', 'URUN', 'PRODUCT', 'ITEM', 'MALZEME', 'HİZMET', 'HIZMET']):
                        val = row[key]
                        urun = str(val).strip() if val is not None else ''
                        break
                
                tutar_raw = 0
                for key in row_keys:
                    key_upper = key.upper().strip()
                    if any(x in key_upper for x in ['TUTAR', 'AMOUNT', 'PRICE', 'FİYAT', 'FIYAT', 'MIKTAR', 'TOPLAM']):
                        tutar_raw = row[key]
                        break
                
                odeme_sekli = ''
                for key in row_keys:
                    key_upper = key.upper().strip()
                    if any(x in key_upper for x in ['ÖDEME', 'ODEME', 'PAYMENT', 'PAY', 'PLAN']):
                        val = row[key]
                        odeme_sekli = str(val).strip() if val is not None else ''
                        break
                
                # Tutar işleme - daha esnek
                tutar = 0
                try:
                    if tutar_raw is not None and tutar_raw != '':
                        if isinstance(tutar_raw, (int, float)):
                            tutar = float(tutar_raw)
                        else:
                            # String işleme - Türkçe sayı formatını destekle
                            tutar_str = str(tutar_raw).strip()
                            if tutar_str:
                                # Virgül ve nokta işleme
                                if ',' in tutar_str and '.' in tutar_str:
                                    # 1.234,56 formatı
                                    tutar_str = tutar_str.replace('.', '').replace(',', '.')
                                elif ',' in tutar_str:
                                    # 1234,56 formatı
                                    tutar_str = tutar_str.replace(',', '.')
                                # Sadece sayı ve nokta bırak
                                import re
                                tutar_str = re.sub(r'[^\d.]', '', tutar_str)
                                tutar = float(tutar_str) if tutar_str else 0
                except Exception as e:
                    print(f"Tutar parse hatası: {e}, raw: {tutar_raw}")
                    tutar = 0
                
                print(f"İşlenmiş veri: TARİH={tarih}, FATURANO='{faturano}', MÜŞTERİ='{musteri}', ÜRÜN='{urun}', TUTAR={tutar}, ÖDEME='{odeme_sekli}'")
                
                # Daha esnek kontrol - en az bir anlamlı veri olsun
                has_meaningful_data = any([
                    faturano and len(faturano.strip()) > 0,
                    musteri and len(musteri.strip()) > 0,
                    urun and len(urun.strip()) > 0,
                    tutar > 0,
                    odeme_sekli and len(odeme_sekli.strip()) > 0
                ])
                
                if not has_meaningful_data:
                    print(f"Satır {i+1} anlamlı veri yok, atlanıyor")
                    continue
                
                # Kayıt oluştur - boş alanları varsayılan değerlerle doldur
                try:
                    hareket = MalzemeHareketi.objects.create(
                        dosya=dosya,
                        tarih=tarih,
                        faturano=(faturano or f'AUTO-{i+1}')[:100],  # Boşsa otomatik numara
                        musteri=(musteri or 'Belirtilmemiş')[:255],
                        urun=(urun or 'Belirtilmemiş')[:255],
                        tutar=tutar,
                        odeme_sekli=(odeme_sekli or 'Belirtilmemiş')[:100],
                        kullanici=request.user,
                    )
                    print(f"Kayıt başarılı: ID={hareket.id}")
                    eklenen += 1
                except Exception as db_error:
                    error_msg = f"Satır {i+1} DB hatası: {db_error}"
                    print(error_msg)
                    hatalar.append(error_msg)
                    continue
                
            except Exception as e:
                error_msg = f"Satır {i+1} işlem hatası: {e}"
                print(error_msg)
                hatalar.append(error_msg)
                continue
        print(f"Toplam eklenen kayıt: {eklenen}")
        print(f"Toplam hata sayısı: {len(hatalar)}")
        
        if eklenen == 0:
            error_detail = f"Hiçbir kayıt eklenemedi. Toplam {len(rows)} satır işlendi."
            if hatalar:
                error_detail += f" İlk 3 hata: {'; '.join(hatalar[:3])}"
            return JsonResponse({'success': False, 'error': error_detail}, status=400)
        
        # Kısmi başarı durumu
        result = {'success': True, 'count': eklenen}
        if hatalar:
            result['warnings'] = f"{len(hatalar)} satırda hata oluştu"
            result['errors'] = hatalar[:5]  # İlk 5 hatayı göster
        
        return JsonResponse(result)
    except Exception as e:
        print(f"Genel hata: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': f'Sunucu hatası: {str(e)}'}, status=500)

def health_check(request):
    """Railway health check endpoint"""
    return JsonResponse({'status': 'healthy', 'timestamp': datetime.now().isoformat()})